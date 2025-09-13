from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from decimal import Decimal
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from .utils import generate_loan_number
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side,NamedStyle
from django.template.loader import get_template
import io
from .models import Borrower, Loan, Collection, ActivityLog
DEFAULT_STATUS = Loan.ACTIVE  # العرض الافتراضي: مستمر
def _log(request, action, target_model="", target_id="", payload=None):
    ActivityLog.objects.create(
        actor=getattr(request, "user", None),
        action=action,
        target_model=target_model,
        target_id=str(target_id),
        payload=payload or {},
    )

def loan_list(request):
    # فلاتر الحالة/نوع السداد/نوع المقترض + بحث بالرقم القومي
    status = request.GET.get("status", DEFAULT_STATUS)   # مستمر افتراضيًا
    repayment = request.GET.get("repayment", "")         # الكل
    btype = request.GET.get("btype", "")                 # الكل
    nid = request.GET.get("nid", "").strip()

    qs = Loan.objects.select_related("borrower").all()

    if status:
        qs = qs.filter(status=status)
    if repayment:
        qs = qs.filter(repayment_type=repayment)
    if btype:
        qs = qs.filter(borrower__borrower_type=btype)
    if nid:
        qs = qs.filter(borrower__national_id__icontains=nid)

    context = {
        "loans": qs.order_by("-received_at")[:500],  # حد مبدئي
        "status": status,
        "repayment": repayment,
        "btype": btype,
        "nid": nid,
    }
    return render(request, "loans/loan_list.html", context)

def loan_add(request):
    if request.method == "POST":
        data = request.POST
        nid = data.get("national_id","").strip()

        # 🔴 شرط الرقم القومي: لازم يبقى 14 رقم
        if not nid.isdigit() or len(nid) != 14:
            return render(request, "loans/loan_add.html", {
                "error": "الرقم القومي يجب أن يكون 14 رقم.",
                "form_data": request.POST
            })
        full_name = data.get("full_name","").strip()
        phone = data.get("phone","").strip()
        address = data.get("address","").strip()
        borrower_type = data.get("borrower_type","employee")

        loan_number = data.get("loan_number","").strip()
        amount = Decimal(data.get("amount","0") or "0")
        repayment_type = data.get("repayment_type","monthly")
        monthly_installment = data.get("monthly_installment") or None
        maturity_date = data.get("maturity_date") or None

        borrower, _ = Borrower.objects.get_or_create(
            national_id=nid,
            defaults=dict(full_name=full_name, phone=phone, address=address, borrower_type=borrower_type),
        )
        borrower.full_name = full_name or borrower.full_name
        borrower.phone = phone or borrower.phone
        borrower.address = address or borrower.address
        borrower.borrower_type = borrower_type or borrower.borrower_type
        borrower.save()

        loan_number = generate_loan_number()
        loan = Loan.objects.create(
            loan_number=loan_number,
            borrower=borrower,
            amount=amount,
            repayment_type=repayment_type,
            monthly_installment=monthly_installment,
            maturity_date=maturity_date or None,
            received_at=timezone.now(),
            total_remaining=amount,
        )
        messages.success(request, "تم إضافة القرض بنجاح.")
        return render(request, "loans/loan_add.html", {
            "success": True
        })
        _log(request, "ADD_LOAN", "Loan", loan.loan_number, payload={"amount": str(amount)})
    return render(request, "loans/loan_add.html", {})
def collect_payment(request, loan_number):
    loan = get_object_or_404(Loan, loan_number=loan_number)

    if request.method == "POST":
        try:
            amt = Decimal(request.POST.get("amount", "0") or "0")
        except Exception:
            messages.error(request, "قيمة غير صالحة.")
            return redirect("loans:loan_list")

        # اجمالي القرض + المدفوع + المتبقي
        amount_total = Decimal(loan.amount or 0)
        total_paid   = Decimal(loan.total_paid or 0)
        remaining    = Decimal(
            loan.total_remaining if loan.total_remaining is not None
            else (amount_total - total_paid)
        )

        # تحقق من صحة القيمة
        if amt <= 0:
            messages.error(request, "المبلغ يجب أن يكون أكبر من صفر.")
            return redirect("loans:loan_list")

        if amt > remaining:
            messages.error(request, f"المبلغ المدخل ({amt}) أكبر من المتبقي ({remaining}).")
            return redirect("loans:loan_list")

        if amt > amount_total:
            messages.error(request, f"المبلغ المدخل ({amt}) أكبر من إجمالي القرض ({amount_total}).")
            return redirect("loans:loan_list")

        # إنشاء عملية التحصيل
        Collection.objects.create(loan=loan, amount=amt, collected_at=timezone.now())
        loan.last_collect_amount = amt
        loan.last_collection_at = timezone.now()
        loan.save(update_fields=['last_collect_amount', 'last_collection_at'])

        _log(request, "COLLECT", "Loan", loan.loan_number, payload={"amount": str(amt)})
        #messages.success(request, f"تم تحصيل مبلغ {amt}.")
        return redirect("loans:loan_list")

    # عرض الفورم
    return render(request, "loans/collect_payment.html", {"loan": loan})

def export_loans_xlsx(request):
    # تصدير البيانات الحالية بعد الفلاتر (نفس منطق loan_list)
    # TODO: استخدم openpyxl/xlsxwriter للتصدير الفعلي
    _log(request, "EXPORT_XLSX", "LoanList")
    return HttpResponse("XLSX export placeholder", content_type="text/plain")

# views.py
from django.db.models import Sum, Max
from django.db.models.functions import Coalesce

# views.py
from decimal import Decimal
from django.db.models import Sum, Max, F, OuterRef, Subquery, Value, DecimalField
from django.db.models.functions import Coalesce
from .models import Borrower, Loan, Collection  # تأكد من الاستيراد

def inquiry(request):
    nid = (request.GET.get("nid") or "").strip()

    borrower = None
    loans = Loan.objects.none()
    agg = {"count": 0, "total": Decimal("0"), "paid": Decimal("0"),
           "remaining": Decimal("0"), "latest_collection_at": None}

    if nid:
        try:
            borrower = Borrower.objects.get(national_id=nid)

            # Subquery للحصول على آخر مبلغ من جدول التحصيلات
            latest_amount_sq = Collection.objects.filter(
                loan=OuterRef('pk')
            ).order_by('-collected_at').values('amount')[:1]

            loans = (
                Loan.objects.filter(borrower=borrower)
                .annotate(
                    last_collect_amount_anno=Coalesce(
                        F('last_collect_amount'),                                   # لو متخزن
                        Subquery(latest_amount_sq, output_field=DecimalField(max_digits=12, decimal_places=2)),  # آخر مبلغ فعلي
                        Value(Decimal('0.00'), output_field=DecimalField(max_digits=12, decimal_places=2)),      # fallback
                    )
                )
                .order_by('-received_at')
            )

            sums = loans.aggregate(
                total=Sum("amount", default=Decimal("0")),
                paid=Sum("total_paid", default=Decimal("0")),
                remaining=Sum("total_remaining", default=Decimal("0")),
                latest_collection_at=Max("last_collection_at"),
            )

            agg.update({
                "count": loans.count(),
                "total": sums["total"],
                "paid": sums["paid"],
                "remaining": sums["remaining"],
                "latest_collection_at": sums["latest_collection_at"],
            })
        except Borrower.DoesNotExist:
            borrower = None

    return render(request, "loans/inquiry.html", {
        "nid": nid,
        "borrower": borrower,
        "loans": loans,
        "agg": agg,
    })

def inquiry_export_pdf(request):
    # TODO: استخدم weasyprint/xhtml2pdf لتوليد PDF من شاشة الاستعلام
    _log(request, "EXPORT_PDF", "Inquiry")
    return HttpResponse("PDF export placeholder", content_type="text/plain")

def logs_readonly(request):
    logs = ActivityLog.objects.all()[:500]
    return render(request, "loans/logs.html", {"logs": logs})

def prefill_by_national_id(request):
    # Endpoint يرجع بيانات مقترض (أو موظف) بالرقم القومي لتعبئة الفورم
    nid = request.GET.get("nid", "").strip()

    # 🔴 شرط الرقم القومي
    if not nid.isdigit() or len(nid) != 14:
        return JsonResponse({"osk": False, "message": "الرقم القومي يجب أن يكون 14 رقم."}, status=400)

    data = {}
    b = Borrower.objects.filter(national_id=nid).first()
    if b:
        data = {
            "full_name": b.full_name,
            "phone": b.phone,
            "address": b.address,
            "borrower_type": b.borrower_type,
        }
    return JsonResponse({"ok": True, **data})
from django.http import HttpResponse
from django.utils import timezone
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import datetime

def export_loans_xlsx(request):
    # فعّل التوقيت المحلي
    timezone.activate(settings.TIME_ZONE)

    # helpers للتاريخ
    def to_excel_dt(dt):
        """
        يقبل datetime أو date:
        - لو date: يدمجه مع 00:00 ويخليه aware في التوقيت المحلي.
        - لو datetime: يضمن إنه aware ثم يحوّله للتوقيت المحلي.
        يرجّع قيمة datetime *naive* مناسبة للإكسيل.
        """
        if not dt:
            return None

        local_tz = timezone.get_default_timezone()

        # حالة DateField: datetime.date بدون وقت
        if isinstance(dt, datetime.date) and not isinstance(dt, datetime.datetime):
            dt = datetime.datetime.combine(dt, datetime.time(0, 0))  # 00:00
            dt = timezone.make_aware(dt, local_tz)

        # حالة datetime
        elif isinstance(dt, datetime.datetime):
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt, local_tz)
        else:
            # أنواع غير متوقعة: رجّع None أو حوّلها لنص حسب رغبتك
            return None

        dt_local = timezone.localtime(dt)
        return dt_local.replace(tzinfo=None)

    def write_dt_cell(cell, dt, *, date_only=False):
        """
        يكتب التاريخ في الخلية بقيمة datetime فعلية (مش نص) ويضبط الفورمات.
        date_only=True يخلي الفورمات يوم-شهر-سنة بدون وقت.
        """
        val = to_excel_dt(dt)
        if val is None:
            cell.value = "-"
        else:
            cell.value = val
            cell.number_format = "yyyy-mm-dd" if date_only else "yyyy-mm-dd hh:mm"

    # نفس فلاتر صفحة القائمة
    loans = Loan.objects.select_related("borrower").all()
    status = request.GET.get("status") or ""
    repayment = request.GET.get("repayment") or ""
    btype = request.GET.get("btype") or ""
    nid = request.GET.get("nid") or ""

    if status:
        loans = loans.filter(status=status)
    if repayment:
        loans = loans.filter(repayment_type=repayment)
    if btype:
        loans = loans.filter(borrower__borrower_type=btype)
    if nid:
        loans = loans.filter(borrower__national_id__icontains=nid)

    # إنشاء الملف
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loans"
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    # تنسيقات
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="EEECE1", end_color="EEECE1", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "الاسم رباعي",
        "العنوان",
        "الرقم القومي",
        "رقم التليفون",
        "قيمة القرض",
        "الحالة",
        "نوع السداد",
        "رقم القرض",
        "تاريخ الاستلام",
        "نوع المقترض",
        "إجمالي المسدد",
        "إجمالي المتبقي",
        "آخر مبلغ تحصيل",
        "آخر تحصيل",
        "تاريخ الاستحقاق",
        "القسط الشهري",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    def status_ar(v): return {"active": "مستمر", "closed": "مكتمل", "bad_debt": "ديون معدومة"}.get(v, v or "-")
    def repay_ar(v): return {"monthly": "أقساط شهرية", "oneoff": "دفعة واحدة"}.get(v, v or "-")
    def btype_ar(v): return {"employee": "ضمن الموظفين", "external": "خارج الموظفين"}.get(v, v or "-")

    # البيانات
    for loan in loans:
        row = [
            getattr(loan.borrower, "full_name", "") or "-",
            getattr(loan.borrower, "address", "") or "-",
            getattr(loan.borrower, "national_id", "") or "-",
            getattr(loan.borrower, "phone", "") or "-",
            float(loan.amount or 0),
            status_ar(loan.status),
            repay_ar(loan.repayment_type),
            loan.loan_number,
            None,
            btype_ar(getattr(loan.borrower, "borrower_type", "")),
            float(loan.total_paid or 0),
            float(loan.total_remaining or 0),
            float(getattr(loan, "last_collect_amount", 0) or 0),
            None,
            None,
            float(loan.monthly_installment or 0),
        ]
        ws.append(row)
        last_row = ws.max_row
        # (9) تاريخ الاستلام: غالبًا datetime — اعرض تاريخ+وقت
        write_dt_cell(ws.cell(row=last_row, column=9), loan.received_at, date_only=False)

        # (14) آخر تحصيل: datetime من الحقل last_collection_at
        write_dt_cell(ws.cell(row=last_row, column=14), getattr(loan, "last_collection_at", None), date_only=False)

        # (15) تاريخ الاستحقاق: DateField — اعرض تاريخ فقط
        write_dt_cell(ws.cell(row=last_row, column=15), loan.maturity_date, date_only=True)
    # بعد إدخال الصفوف: عيّن فورمات التواريخ
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for idx, cell in enumerate(row, start=1):
            cell.alignment = center
            cell.border = border
            if idx in (5, 11, 12, 13, 16) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            if idx in (9, 14, 15) and isinstance(cell.value, datetime.datetime):
                # idx 9: تاريخ الاستلام, idx 14: آخر تحصيل, idx 15: تاريخ الاستحقاق
                cell.number_format = "yyyy-mm-dd hh:mm" if idx in (9, 14) else "yyyy-mm-dd"

    widths = [18, 22, 16, 14, 12, 12, 14, 18, 18, 14, 14, 14, 14, 18, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="loans.xlsx"'
    wb.save(resp)
    return resp


# imports المطلوبة
from django.http import HttpResponse, HttpResponseBadRequest
from django.db.models import OuterRef, Subquery, DateTimeField, DecimalField, Value
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.conf import settings
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

# ... imports موديلاتك
# from .models import Borrower, Loan, Collection

def inquiry_export_xlsx(request):
    # فعِّل التوقيت المحلي داخل الفيو (مش على مستوى الملف)
    timezone.activate(settings.TIME_ZONE)

    # هيلبرز للتواريخ مع الإكسيل
    def to_excel_dt(dt):
        if not dt:
            return None
        if timezone.is_naive(dt):
            # عامله كـ local aware لو جالك Naive
            dt = timezone.make_aware(dt, timezone.get_default_timezone())
        dt_local = timezone.localtime(dt)
        return dt_local.replace(tzinfo=None)

    def write_dt_cell(cell, dt):
        val = to_excel_dt(dt)
        if val is None:
            cell.value = "-"
        else:
            cell.value = val
            cell.number_format = "yyyy-mm-dd hh:mm"

    # ===== بيانات الإدخال =====
    nid = (request.GET.get("nid") or "").strip()
    if not nid:
        # رجّع Response واضح بدل None
        return HttpResponseBadRequest("parameter 'nid' is required")

    borrower = None
    loans = Loan.objects.none()
    agg = {"count": 0, "total": 0, "paid": 0, "remaining": 0}

    # ===== جلب البيانات مع Subquery لآخر تحصيل =====
    try:
        borrower = Borrower.objects.get(national_id=nid)

        latest_dt_sq = Collection.objects.filter(
            loan=OuterRef('pk')
        ).order_by('-collected_at').values('collected_at')[:1]

        latest_amt_sq = Collection.objects.filter(
            loan=OuterRef('pk')
        ).order_by('-collected_at').values('amount')[:1]

        loans = (
            Loan.objects.filter(borrower=borrower)
            .annotate(
                latest_collection_dt=Subquery(latest_dt_sq, output_field=DateTimeField()),
                latest_collection_amount=Coalesce(
                    Subquery(latest_amt_sq, output_field=DecimalField(max_digits=12, decimal_places=2)),
                    Value(Decimal('0.00'), output_field=DecimalField(max_digits=12, decimal_places=2))
                ),
            )
            .order_by('-received_at')
        )

        agg = {
            "count": loans.count(),
            "total": sum(float(l.amount or 0) for l in loans),
            "paid": sum(float(l.total_paid or 0) for l in loans),
            "remaining": sum(float(l.total_remaining or 0) for l in loans),
        }
    except Borrower.DoesNotExist:
        borrower = None
        loans = Loan.objects.none()  # تأكيد

    # ===== بناء ملف الإكسيل =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inquiry"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    title = ws.cell(row=r, column=1, value="تقرير استعلام عن المقترض")
    title.font = Font(bold=True, size=14)
    title.alignment = center
    r += 2

    # بيانات المقترض
    ws.cell(row=r, column=1, value="البيانات الشخصية").font = header_font
    r += 1

    def bt_ar(v):
        return {"employee": "ضمن الموظفين", "external": "خارج الموظفين"}.get(v, v or "-")

    if borrower:
        info = [
            ("الاسم", borrower.full_name or "-"),
            ("الرقم القومي", borrower.national_id or "-"),
            ("العنوان", borrower.address or "-"),
            ("الهاتف", borrower.phone or "-"),
            ("نوع المقترض", bt_ar(getattr(borrower, "borrower_type", ""))),
            ("عدد القروض", agg["count"]),
        ]
        for label, val in info:
            ws.cell(row=r, column=1, value=label).font = header_font
            v = ws.cell(row=r, column=2, value=val)
            v.alignment = right
            r += 1
    else:
        ws.cell(row=r, column=1, value="لا يوجد مقترض بهذا الرقم القومي.").alignment = right
        r += 2

    r += 1

    # الملخص
    ws.cell(row=r, column=1, value="الملخص").font = header_font
    r += 1

    last_dt = None
    if loans:
        last_dt = max((l.latest_collection_dt for l in loans if l.latest_collection_dt), default=None)

    summary = [
        ("إجمالي القروض", agg["total"]),
        ("إجمالي المسدد", agg["paid"]),
        ("إجمالي المتبقي", agg["remaining"]),
        ("آخر تحصيل", last_dt),
    ]
    for label, val in summary:
        ws.cell(row=r, column=1, value=label).font = header_font
        c = ws.cell(row=r, column=2)
        if isinstance(val, datetime.datetime):
            write_dt_cell(c, val)
        else:
            c.value = val
            if isinstance(val, (int, float)):
                c.number_format = "#,##0.00"
        c.alignment = right
        r += 1

    r += 2

    # جدول القروض
    headers = [
        "آخر مبلغ تحصيل",
        "رقم القرض",
        "قيمة القرض",
        "نوع السداد",
        "الحالة",
        "القسط الشهري",
        "المسدد",
        "المتبقي",
        "تاريخ الاستلام",
        "تاريخ الاستحقاق",
        "آخر تحصيل",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    def ar_status(v):   return {"active": "مستمر", "closed": "مكتمل", "bad_debt": "ديون معدومة"}.get(v, v or "-")
    def ar_repay(v):    return {"monthly": "أقساط شهرية", "oneoff": "دفعة واحدة"}.get(v, v or "-")

    for l in loans:
        monthly_val = float(l.monthly_installment or 0) if l.repayment_type == "monthly" else "-"
        last_col_dt = to_excel_dt(l.latest_collection_dt)

        row = [
            float(l.latest_collection_amount or 0),
            l.loan_number,
            float(l.amount or 0),
            ar_repay(l.repayment_type),
            ar_status(l.status),
            monthly_val,
            float(l.total_paid or 0),
            float(l.total_remaining or 0),
            l.received_at.strftime("%Y-%m-%d") if l.received_at else "-",
            l.maturity_date.strftime("%Y-%m-%d") if l.maturity_date else "-",
            last_col_dt or "-",
        ]
        ws.append(row)

        # فورمات التاريخ لعمود "آخر تحصيل"
        last_row = ws.max_row
        cell_dt = ws.cell(row=last_row, column=11)
        if isinstance(cell_dt.value, datetime.datetime):
            cell_dt.number_format = "yyyy-mm-dd hh:mm"
        elif cell_dt.value != "-":
            cell_dt.value = last_col_dt or "-"
            if last_col_dt:
                cell_dt.number_format = "yyyy-mm-dd hh:mm"

    # تنسيقات عامة
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for idx, cell in enumerate(row, start=1):
            cell.alignment = center
            cell.border = border
            if idx in (1, 3, 7, 8) and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
            if idx == 6 and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    widths = [18, 14, 16, 12, 14, 14, 14, 16, 16, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A{}".format(header_row + 1)

    # ===== ارجع Response دايمًا =====
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fname = f"inquiry_{nid or 'unknown'}.xlsx"
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp
