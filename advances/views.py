from datetime import datetime
from io import BytesIO
from decimal import Decimal

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import ValidationError
from django.db.models import Q, Count
from django.http import (
    JsonResponse, HttpResponse, HttpResponseBadRequest
)
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from .models import AdvancePeriod, AdvanceRequest, AdvanceType, AdvanceStatus
from .forms import AdvanceRequestForm


def is_admin(user):
    return user.is_superuser or user.is_staff


# ======================
# Helpers
# ======================

def _allowed_range_for_user(user):
    """يرجع (min, max) أو (None, None) لو الراتب الأساسي غير صالح."""
    base = getattr(user, 'base_salary', None) or Decimal('0')
    try:
        base = Decimal(base)
    except Exception:
        base = Decimal('0')

    if base <= 0:
        return (None, None)

    min_amt = Decimal('100')
    max_amt = (base / Decimal('4')).quantize(Decimal('0.01'))
    return (min_amt, max_amt)


def _set_amount_helptext_on_form(form, user):
    """يضبط help_text لحقل amount لو القالب بيعرضه."""
    if not hasattr(form, 'fields') or 'amount' not in form.fields:
        return
    min_amt, max_amt = _allowed_range_for_user(user)
    if min_amt is None or max_amt is None:
        form.fields['amount'].help_text = "لا يمكنك طلب سلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر."
    else:
        form.fields['amount'].help_text = f"مسموح لك بطلب سلفة من {min_amt} إلى {max_amt}."


# ======================
# واجهة الموظف
# ======================

@login_required
def user_advances_portal(request):
    period_first = AdvancePeriod.objects.filter(
        advance_type=AdvanceType.FIRST, is_active=True
    ).order_by('-start_date').first()
    period_second = AdvancePeriod.objects.filter(
        advance_type=AdvanceType.SECOND, is_active=True
    ).order_by('-start_date').first()

    can_first = period_first.is_open_now() if period_first else False
    can_second = period_second.is_open_now() if period_second else False

    base = getattr(request.user, 'base_salary', None) or Decimal('0')
    try:
        base = Decimal(base)
    except Exception:
        base = Decimal('0')
    salary_zero = (base <= 0)

    # ✅ فلاغ للسماح/المنع
    can_request = True
    reason = ""
    if salary_zero:
        can_request = False
        reason = "غير مسموح بطلب سلفة لأن راتبك الأساسي غير مُسجّل أو يساوي صفر."
    elif not (can_first or can_second):
        can_request = False
        reason = "لا توجد فترة سُلفة مفتوحة حالياً."

    # ✅ الفترات الحالية
    first_has_current = AdvanceRequest.objects.filter(
        user=request.user, advance_type=AdvanceType.FIRST, period=period_first
    ).exists() if period_first else False

    second_has_current = AdvanceRequest.objects.filter(
        user=request.user, advance_type=AdvanceType.SECOND, period=period_second
    ).exists() if period_second else False

    # الفورم دايمًا موجود
    available_types = []
    if can_first:
        available_types.append(AdvanceType.FIRST)
    if can_second:
        available_types.append(AdvanceType.SECOND)

    form = AdvanceRequestForm(available_types=available_types)
    _set_amount_helptext_on_form(form, request.user)

    hint_min, hint_max = _allowed_range_for_user(request.user)
    hint_text = (
        "لا يمكنك طلب سلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر."
        if hint_min is None else f"مسموح لك بطلب سلفة من {hint_min} إلى {hint_max}."
    )

    user_reqs = AdvanceRequest.objects.filter(
        user=request.user
    ).select_related('period').order_by('-created_at')

    first_latest = AdvanceRequest.objects.filter(
        user=request.user, advance_type=AdvanceType.FIRST, period=period_first
    ).order_by('-created_at').first() if period_first else None

    second_latest = AdvanceRequest.objects.filter(
        user=request.user, advance_type=AdvanceType.SECOND, period=period_second
    ).order_by('-created_at').first() if period_second else None

    ctx = {
        'form': form,
        'can_first': can_first,
        'can_second': can_second,
        'period_first': period_first,
        'period_second': period_second,
        'first_latest': first_latest,
        'second_latest': second_latest,
        'user_reqs': user_reqs,
        'first_has_current': first_has_current,
        'second_has_current': second_has_current,
        'salary_zero': salary_zero,
        'amount_hint_min': hint_min,
        'amount_hint_max': hint_max,
        'amount_hint_text': hint_text,
        'can_request': can_request,
        'reason': reason,
    }
    return render(request, 'advances/my_advances.html', ctx)

@login_required
def submit_advance(request):
    """إنشاء طلب الموظف."""
    if request.method != 'POST':
        return HttpResponseBadRequest('Invalid method')

    advance_type = request.POST.get('advance_type')
    period = AdvancePeriod.objects.filter(
        advance_type=advance_type, is_active=True
    ).order_by('-start_date').first()
    if not period or not period.is_open_now():
        return JsonResponse({'ok': False, 'msg': 'الفترة غير متاحة الآن لهذا النوع'})

    # منع الإرسال لو الراتب = 0
    min_amt, max_amt = _allowed_range_for_user(request.user)
    if min_amt is None:  # يعني الراتب الأساسي <= 0
        return JsonResponse({'ok': False, 'msg': 'غير مسموح بطلب سلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر.'})

    # منع الإرسال لو فيه طلب قائم لهذا النوع في نفس الفترة
    existing = AdvanceRequest.objects.filter(
        user=request.user, advance_type=advance_type, period=period
    ).first()
    if existing:
        return JsonResponse({'ok': False, 'msg': 'لديك طلب قائم لهذا النوع في الفترة الحالية. استخدم زر "تعديل" أو "حذف".'})

    form = AdvanceRequestForm(request.POST, available_types=[advance_type])
    form.instance.user = request.user

    if form.is_valid():
        obj = form.save(commit=False)
        obj.user = request.user
        obj.period = period
        obj.status = AdvanceStatus.UNDER_REVIEW
        obj.locked = False
        try:
            obj.save()
        except ValidationError:
            return JsonResponse({'ok': False, 'msg': f'السلفة يجب أن تكون بين {min_amt} و {max_amt}.'})
        return JsonResponse({'ok': True})
    else:
        return JsonResponse({'ok': False, 'msg': f'السلفة يجب أن تكون بين {min_amt} و {max_amt}.'})


@login_required
def user_edit_advance(request, pk: int):
    """تعديل الموظف لطلبه طالما 'تحت المراجعة'."""
    obj = get_object_or_404(AdvanceRequest, pk=pk, user=request.user)

    if request.method != 'POST':
        return HttpResponseBadRequest('Invalid method')

    if obj.status != AdvanceStatus.UNDER_REVIEW or obj.locked or obj.admin_decision or obj.user_locked:
        return JsonResponse({'ok': False, 'msg': 'غير مسموح بتعديل هذا الطلب الآن.'}, status=403)

    amount_raw = (request.POST.get('amount') or '').strip()

    try:
        obj.amount = Decimal(amount_raw)
    except Exception:
        min_amt, max_amt = _allowed_range_for_user(request.user)
        hint = ("لا يمكنك طلب سلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر."
                if min_amt is None else f"المسموح من {min_amt} إلى {max_amt}.")
        return JsonResponse({'ok': False, 'msg': f'مبلغ غير صالح. {hint}'})

    try:
        obj.save()
    except ValidationError:
        min_amt, max_amt = _allowed_range_for_user(request.user)
        err = ("غير مسموح بطلب سلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر."
               if min_amt is None else f"السلفة يجب أن تكون بين {min_amt} و {max_amt}.")
        return JsonResponse({'ok': False, 'msg': err})

    return JsonResponse({'ok': True, 'amount': str(obj.amount)})


@login_required
def user_delete_advance(request, pk: int):
    """حذف طلب الموظف (قبل التأكيد النهائي فقط) — يرجّع JSON."""
    obj = get_object_or_404(AdvanceRequest, pk=pk, user=request.user)
    if request.method != 'POST':
        return HttpResponseBadRequest('Invalid method')

    if obj.status != AdvanceStatus.UNDER_REVIEW or obj.locked:
        return JsonResponse({'ok': False, 'msg': 'لا يمكن حذف هذا الطلب الآن.'}, status=403)

    if obj.admin_decision:
        return JsonResponse({'ok': False, 'msg': 'لا يمكن حذف الطلب بعد وجود قرار مبدئي.'}, status=403)

    obj.delete()
    return JsonResponse({'ok': True})


# ======================
# إدارة مواعيد السلف (أدمن)
# ======================

@user_passes_test(is_admin)
def periods_manage(request):
    """إدارة ميعاد السُلفة: سجل واحد فعلي لكل نوع."""
    def get_single(type_code):
        qs = AdvancePeriod.objects.filter(advance_type=type_code).order_by('-is_active', '-start_date', '-id')
        obj = qs.first()
        if obj is None:
            obj = AdvancePeriod.objects.create(
                advance_type=type_code,
                start_date=timezone.localdate(),
                end_date=timezone.localdate(),
                is_active=False,
            )
        qs.exclude(pk=obj.pk).update(is_active=False)
        return obj

    first_period = get_single(AdvanceType.FIRST)
    second_period = get_single(AdvanceType.SECOND)

    if request.method == 'POST':
        which = request.POST.get('which')          # 'FIRST' or 'SECOND'
        s = request.POST.get('start_date')         # 'YYYY-MM-DD'
        e = request.POST.get('end_date')           # 'YYYY-MM-DD'
        active = request.POST.get('is_active') in ('on', 'true', '1')

        try:
            s_date = datetime.strptime(s, '%Y-%m-%d').date()
            e_date = datetime.strptime(e, '%Y-%m-%d').date()
        except Exception:
            return redirect('adv-periods')

        p = first_period if which == AdvanceType.FIRST else second_period
        p.start_date = s_date
        p.end_date = e_date
        p.is_active = active
        p.save()

        AdvancePeriod.objects.filter(advance_type=which).exclude(pk=p.pk).update(is_active=False)

        return redirect('adv-periods')

    return render(request, 'advances/periods_manage.html', {
        'first_period': first_period,
        'second_period': second_period,
    })


# ======================
# إدارة الطلبات (أدمن)
# ======================

@user_passes_test(is_admin)
def requests_list(request):
    """قائمة الطلبات + فلاتر + بحث بالاسم"""
    q_status = request.GET.get('status')           # UNDER_REVIEW / APPROVED / REJECTED
    q_type   = request.GET.get('type')             # FIRST / SECOND
    q_complete = request.GET.get('complete')       # yes/no
    q_cycle = request.GET.get('cycle')             # complete/incomplete
    q = request.GET.get('q')                       # نص البحث

    qs = AdvanceRequest.objects.select_related('user', 'period')

    if q_status in dict(AdvanceStatus.choices):
        qs = qs.filter(status=q_status)
    if q_type in dict(AdvanceType.choices):
        qs = qs.filter(advance_type=q_type)
    if q_complete == 'yes':
        qs = qs.exclude(status=AdvanceStatus.UNDER_REVIEW)
    elif q_complete == 'no':
        qs = qs.filter(status=AdvanceStatus.UNDER_REVIEW)

    if q:
        qs = qs.filter(
            Q(user__username__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q)
        )

    if q_cycle in ('complete', 'incomplete'):
        active_periods = AdvancePeriod.objects.filter(is_active=True)
        users_done = (
            AdvanceRequest.objects.filter(period__in=active_periods)
            .exclude(status=AdvanceStatus.UNDER_REVIEW)
            .values('user')
            .annotate(done_types=Count('advance_type', filter=~Q(status=AdvanceStatus.UNDER_REVIEW), distinct=True))
            .filter(done_types=2)
            .values_list('user', flat=True)
        )
        if q_cycle == 'complete':
            qs = qs.filter(user_id__in=list(users_done))
        else:
            qs = qs.exclude(user_id__in=list(users_done))

    qs = qs.order_by('-created_at')
    return render(request, 'advances/requests_list.html', {'requests': qs})


# --- قرارات مبدئية (فردي) ---
@user_passes_test(is_admin)
def approve_one(request, pk):
    """تعيين قرار مبدئي (قبول) لطلب واحد فقط + قفل تعديل الموظف"""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'msg': 'Invalid method'})
    obj = get_object_or_404(AdvanceRequest, pk=pk)
    obj.admin_decision = AdvanceStatus.APPROVED
    obj.user_locked = True
    obj.save(update_fields=['admin_decision', 'user_locked'])
    return JsonResponse({
        'ok': True,
        'msg': 'تم تسجيل قرار مبدئي: قبول',
        'admin_decision_code': 'APPROVED',
        'admin_decision_display': obj.get_admin_decision_display(),
    })


@user_passes_test(is_admin)
def reject_one(request, pk):
    """تعيين قرار مبدئي (رفض) لطلب واحد فقط + قفل تعديل الموظف"""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'msg': 'Invalid method'})
    obj = get_object_or_404(AdvanceRequest, pk=pk)
    obj.admin_decision = AdvanceStatus.REJECTED
    obj.user_locked = True
    obj.save(update_fields=['admin_decision', 'user_locked'])
    return JsonResponse({
        'ok': True,
        'msg': 'تم تسجيل قرار مبدئي: رفض',
        'admin_decision_code': 'REJECTED',
        'admin_decision_display': obj.get_admin_decision_display(),
    })


# --- قرارات مبدئية (باقي الطلبات) ---
@user_passes_test(is_admin)
def approve_rest(request):
    """تعيين قرار مبدئي (قبول) لكل الطلبات التي لا يوجد لها قرار مبدئي حتى الآن"""
    n = AdvanceRequest.objects.filter(
        status=AdvanceStatus.UNDER_REVIEW,
        admin_decision__isnull=True
    ).update(admin_decision=AdvanceStatus.APPROVED, user_locked=True)
    return JsonResponse({'ok': True, 'count': n})


@user_passes_test(is_admin)
def reject_rest(request):
    """تعيين قرار مبدئي (رفض) لكل الطلبات التي لا يوجد لها قرار مبدئي حتى الآن"""
    n = AdvanceRequest.objects.filter(
        status=AdvanceStatus.UNDER_REVIEW,
        admin_decision__isnull=True
    ).update(admin_decision=AdvanceStatus.REJECTED, user_locked=True)
    return JsonResponse({'ok': True, 'count': n})


# --- تأكيد نهائي ---
@user_passes_test(is_admin)
def confirm_when_no_under_review(request):
    """
    تأكيد نهائي: لا يسمح بالتأكيد طالما هناك طلبات UNDER_REVIEW بلا قرار مبدئي.
    عند التأكيد: ننقل admin_decision إلى status ونقفل الطلبات (locked=True).
    """
    has_undecided = AdvanceRequest.objects.filter(
        status=AdvanceStatus.UNDER_REVIEW,
        admin_decision__isnull=True
    ).exists()
    if has_undecided:
        return JsonResponse({'ok': False, 'msg': 'لا يمكن التأكيد: توجد طلبات بلا قرار مبدئي.'})

    qs_to_finalize = AdvanceRequest.objects.filter(
        status=AdvanceStatus.UNDER_REVIEW,
        admin_decision__in=[AdvanceStatus.APPROVED, AdvanceStatus.REJECTED]
    )
    updated = 0
    for r in qs_to_finalize:
        r.status = r.admin_decision
        r.locked = True
        r.save(update_fields=['status', 'locked'])
        updated += 1

    return JsonResponse({'ok': True, 'msg': f'تم التأكيد النهائي وقفل الطلبات ({updated}).'})


# --- تعديل الأدمن للطلب (قبل التأكيد النهائي فقط) ---
@user_passes_test(is_admin)
def admin_edit_request(request, pk: int):
    """
    تعديل الأدمن للطلب: amount, notes, decision (APPROVED/REJECTED/UNDER_REVIEW)
    """
    obj = get_object_or_404(AdvanceRequest, pk=pk)

    if obj.locked:
        return render(request, 'advances/admin_request_edit.html', {
            'obj': obj,
            'err': 'تم تأكيد هذا الطلب ولا يمكن تعديله.'
        })

    if request.method == 'POST':
        amount_raw = request.POST.get('amount', '').strip()
        notes = request.POST.get('notes', '')
        decision = (request.POST.get('decision') or request.POST.get('status') or '').strip()

        try:
            obj.amount = float(amount_raw) if amount_raw != '' else obj.amount
        except Exception:
            min_amt, max_amt = _allowed_range_for_user(obj.user)
            hint = (
                "لا يمكن تعديل السلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر."
                if min_amt is None else f"المسموح من {min_amt} إلى {max_amt}."
            )
            return render(request, 'advances/admin_request_edit.html', {
                'obj': obj,
                'err': f'مبلغ غير صالح. {hint}'
            })

        obj.notes = notes

        if decision in (AdvanceStatus.APPROVED, AdvanceStatus.REJECTED):
            obj.admin_decision = decision
            obj.user_locked = True
        else:
            obj.admin_decision = None
            obj.user_locked = False

        try:
            obj.save(update_fields=['amount', 'notes', 'admin_decision', 'user_locked', 'updated_at'])
        except ValidationError:
            min_amt, max_amt = _allowed_range_for_user(obj.user)
            err = (
                "غير مسموح بالسلفة لأن الراتب الأساسي غير مُسجّل أو يساوي صفر."
                if min_amt is None else f"السلفة يجب أن تكون بين {min_amt} و {max_amt}."
            )
            return render(request, 'advances/admin_request_edit.html', {'obj': obj, 'err': err})

        return redirect('adv-requests')

    return render(request, 'advances/admin_request_edit.html', {'obj': obj})


# --- تصدير إكسل ---
@user_passes_test(is_admin)
def export_requests_xlsx(request):
    """تصدير Excel — بدون Notes + مع (اسم الفرع، رقم الحساب)."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("Package 'openpyxl' مطلوب: نفّذ pip install openpyxl", status=500)

    q_status = request.GET.get('status')  # APPROVED / REJECTED
    q_type = request.GET.get('type')      # FIRST / SECOND

    qs = AdvanceRequest.objects.select_related('user', 'period')
    if q_status in dict(AdvanceStatus.choices):
        qs = qs.filter(status=q_status)
    if q_type in dict(AdvanceType.choices):
        qs = qs.filter(advance_type=q_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "Advances"

    headers = [
        'Employee', 'اسم الفرع', 'رقم الحساب',
        'Type', 'Amount', 'Final Status', 'Proposed (Admin)',
        'Period Start', 'Period End', 'Submitted At'
    ]
    ws.append(headers)

    for r in qs:
        u = r.user
        branch = getattr(u, 'branch_name', '') or ''
        bank   = getattr(u, 'bank_account_number', '') or ''

        ws.append([
            getattr(u, 'username', u.id),
            branch,
            bank,
            dict(AdvanceType.choices)[r.advance_type],
            float(r.amount),
            dict(AdvanceStatus.choices)[r.status],
            dict(AdvanceStatus.choices).get(r.admin_decision, '') if r.admin_decision else '',
            str(r.period.start_date),
            str(r.period.end_date),
            r.created_at.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M'),
        ])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.column_dimensions[get_column_letter(headers.index('رقم الحساب') + 1)].width = 22

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"advances_{q_type or 'ALL'}_{q_status or 'ALL'}.xlsx"
    resp = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# --- عمليات شهرية (اختياري كما هو) ---
@user_passes_test(is_admin)
def full_month_reset(request):
    AdvanceRequest.objects.all().delete()
    AdvancePeriod.objects.update(is_active=False)
    return JsonResponse({'ok': True, 'msg': 'تم الريست الشهري وحُذفت كل الطلبات وتم تعطيل الفترات'})


@user_passes_test(is_admin)
def delete_first_advance_requests(request):
    n, _ = AdvanceRequest.objects.filter(advance_type=AdvanceType.FIRST).delete()
    return JsonResponse({'ok': True, 'count': n})

@user_passes_test(is_admin)
def sync_push(request):
    return JsonResponse({'ok': True, 'msg': 'تم التحديث للمستخدمين'})
#--------------------------------------api-----------------------------
from rest_framework import viewsets, permissions, status
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import AdvanceRequest, AdvancePeriod
from .serializers import AdvanceRequestSerializer
from decimal import Decimal
from django.core.exceptions import ValidationError as DjangoValidationError

def format_error_messages(e):
    messages = []
    if hasattr(e, "message_dict"):
        for _, errs in e.message_dict.items():
            messages.extend(errs)
    elif hasattr(e, "messages"):
        messages.extend(e.messages)
    else:
        messages = [str(e)]
    return messages


class AdvanceRequestViewSet(viewsets.ModelViewSet):
    serializer_class = AdvanceRequestSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = AdvanceRequest.objects.none()  # placeholder عشان الـ router

    def get_queryset(self):
        # يعرض طلبات الموظف الحالي بس
        return AdvanceRequest.objects.filter(user=self.request.user)

    def create(self, request, *args, **kwargs):
        """تسجيل طلب سلفة جديد"""
        try:
            response = super().create(request, *args, **kwargs)
            return Response({
                "success": True,
                "message": "تم تسجيل طلب السلفة بنجاح.",
                "data": response.data
            }, status=200)
        except (DjangoValidationError, ValidationError, Exception) as e:
            return Response({
                "success": False,
                "message": format_error_messages(e)
            }, status=400)

    def update(self, request, *args, **kwargs):
        """تعديل طلب سلفة"""
        instance = self.get_object()

        # تحقق الصلاحيات
        if instance.user != self.request.user:
            return Response({"success": False, "message": ["لا تملك صلاحية تعديل هذا الطلب."]}, status=400)

        today = timezone.localdate()
        if not (instance.period.start_date <= today <= instance.period.end_date and instance.period.is_active):
            return Response({"success": False, "message": ["لا يمكن تعديل طلب السلفة لأن الفترة مغلقة."]}, status=400)

        if instance.status != "UNDER_REVIEW" or instance.locked or instance.admin_decision or getattr(instance, "user_locked", False):
            return Response({"success": False, "message": ["لا يمكن تعديل طلب السلفة بعد وجود قرار مبدئي أو تأكيد."]}, status=400)

        # محاولة التعديل
        try:
            response = super().update(request, *args, **kwargs)
            return Response({
                "success": True,
                "message": "تم تعديل طلب السلفة بنجاح.",
                "data": response.data
            }, status=200)
        except (DjangoValidationError, ValidationError, Exception) as e:
            return Response({
                "success": False,
                "message": format_error_messages(e)
            }, status=400)

    def destroy(self, request, *args, **kwargs):
        """حذف طلب سلفة"""
        instance = self.get_object()

        if instance.user != self.request.user:
            return Response({"success": False, "message": ["لا تملك صلاحية حذف هذا الطلب."]}, status=400)

        today = timezone.localdate()
        if not (instance.period.start_date <= today <= instance.period.end_date and instance.period.is_active):
            return Response({"success": False, "message": ["لا يمكن حذف طلب السلفة لأن الفترة مغلقة."]}, status=400)

        if instance.status != "UNDER_REVIEW" or instance.locked or instance.admin_decision or getattr(instance, "user_locked", False):
            return Response({"success": False, "message": ["لا يمكن حذف طلب السلفة بعد وجود قرار مبدئي أو تأكيد."]}, status=400)

        instance.delete()
        return Response({"success": True, "message": "تم حذف طلب السلفة بنجاح."}, status=200)

    def perform_create(self, serializer):
        user = self.request.user
        today = timezone.localdate()

        active_period = AdvancePeriod.objects.filter(
            is_active=True,
            start_date__lte=today,
            end_date__gte=today
        ).first()

        if not active_period:
            raise DjangoValidationError(["لا توجد فترة سلف متاحة حالياً."])

        if AdvanceRequest.objects.filter(
            user=user,
            advance_type=active_period.advance_type,
            period=active_period
        ).exists():
            raise DjangoValidationError(["لديك بالفعل طلب سلفة مسجل لهذه الفترة."])

        try:
            serializer.save(
                user=user,
                period=active_period,
                advance_type=active_period.advance_type
            )
        except (DjangoValidationError, ValidationError, Exception) as e:
            # بدال ما يطلع exception، نرجعه كـ response منظم
            raise DjangoValidationError(format_error_messages(e))


#--------------------------------------------------------------
class AdvanceEligibilityView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        today = timezone.localdate()

        # المرتب الأساسي
        base = getattr(user, 'base_salary', None) or Decimal('0')

        # قيم الحدود
        min_val = Decimal('100') if base > 0 else None
        max_val = (Decimal(base) / Decimal('4')).quantize(Decimal('0.01')) if base > 0 else None

        # هل فيه فترة مفتوحة؟
        active_period = AdvancePeriod.objects.filter(
            is_active=True,
            start_date__lte=today,
            end_date__gte=today
        ).first()

        eligible = bool(base > 0 and active_period)

        return Response({
            "success": True,
            "eligible": eligible,
            "min_amount": float(min_val) if min_val else None,
            "max_amount": float(max_val) if max_val else None
        }, status=200)
